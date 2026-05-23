from flask import Flask, render_template, request, redirect, url_for
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

ARCHIVO_EXCEL = "inscripciones_vacaciones.xlsx"

HORARIOS = [
    "Martes 16 de Junio entre las 9:00 a. m. y 12:00 p. m.",
    "Martes 16 de Junio entre las 2:00 p. m. y 5:00 p. m.",
    "Miércoles 17 de Junio entre las 9:00 a. m. y 12:00 p. m.",
    "Miércoles 17 de Junio entre las 2:00 p. m. y 5:00 p. m.",
    "Jueves 18 de Junio entre las 9:00 a. m. y 12:00 p. m.",
    "Jueves 18 de Junio entre las 2:00 p. m. y 5:00 p. m.",
    "Viernes 19 de Junio entre las 9:00 a. m. y 12:00 p. m.",
    "Viernes 19 de Junio entre las 2:00 p. m. y 5:00 p. m."
]

ENCABEZADOS = [
    "ID",
    "Autorización Datos",
    "Autorización Imágenes",
    "Tipo Documento",
    "Número Documento",
    "Nombre Completo",
    "Celular",
    "Correo",
    "Comunidad",
    "Espacio Adecuado",
    "Horario"
]

def crear_excel_si_no_existe():
    if not os.path.exists(ARCHIVO_EXCEL):
        wb = Workbook()
        ws = wb.active
        ws.title = "Inscripciones"
        ws.append(ENCABEZADOS)
        wb.save(ARCHIVO_EXCEL)

def leer_registros():
    crear_excel_si_no_existe()
    wb = load_workbook(ARCHIVO_EXCEL)
    ws = wb["Inscripciones"]

    registros = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if any(fila):
            registros.append(fila)

    return registros

def limpiar_texto(texto):
    if texto is None:
        return ""
    return str(texto).strip()

def obtener_horarios_disponibles():
    crear_excel_si_no_existe()

    wb = load_workbook(ARCHIVO_EXCEL)
    ws = wb["Inscripciones"]

    horarios_ocupados = []

    for fila in ws.iter_rows(min_row=2, values_only=True):
        horario = limpiar_texto(fila[10])
        if horario:
            horarios_ocupados.append(horario)

    wb.close()

    horarios_disponibles = [
        horario for horario in HORARIOS
        if limpiar_texto(horario) not in horarios_ocupados
    ]

    return horarios_disponibles

    return [h for h in HORARIOS if h not in horarios_ocupados]

def guardar_registro(datos):
    crear_excel_si_no_existe()

    wb = load_workbook(ARCHIVO_EXCEL)
    ws = wb["Inscripciones"]

    horario_seleccionado = limpiar_texto(datos[-1])

    horarios_ocupados = []

    for fila in ws.iter_rows(min_row=2, values_only=True):
        horario = limpiar_texto(fila[10])
        if horario:
            horarios_ocupados.append(horario)

    if horario_seleccionado in horarios_ocupados:
        wb.close()
        return False

    nuevo_id = ws.max_row

    datos_limpios = tuple(limpiar_texto(dato) for dato in datos)

    ws.append((nuevo_id,) + datos_limpios)

    wb.save(ARCHIVO_EXCEL)
    wb.close()

    return True

@app.route("/", methods=["GET", "POST"])
def formulario():
    if request.method == "POST":
        datos = (
            request.form.get("autorizacion_datos"),
            request.form.get("autorizacion_imagenes"),
            request.form.get("tipo_documento"),
            request.form.get("numero_documento"),
            request.form.get("nombre_completo"),
            request.form.get("celular"),
            request.form.get("correo"),
            request.form.get("comunidad"),
            request.form.get("espacio_adecuado"),
            request.form.get("horario")
        )

        guardado = guardar_registro(datos)

        if guardado:
            return redirect(url_for("gracias"))
        else:
            return "Ese horario ya fue seleccionado por otra persona. Por favor vuelve al formulario y escoge otro horario disponible."

    horarios_disponibles = obtener_horarios_disponibles()
    return render_template("index.html", horarios=horarios_disponibles)

@app.route("/gracias")
def gracias():
    return render_template("gracias.html")

@app.route("/registros")
def registros():
    datos = leer_registros()
    return render_template("registros.html", datos=datos)

crear_excel_si_no_existe()
if __name__ == "__main__":
    app.run()
